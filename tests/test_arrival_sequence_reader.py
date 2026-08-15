from __future__ import annotations

import os
import tempfile
import unittest
import zipfile
import zlib
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from Modules.Excel.ArrivalSequenceReader import (
    ArrivalSequenceEntry,
    ArrivalSequenceError,
    ArrivalSequenceReader,
    ArrivalSequenceSnapshot,
    ArrivalSummary,
    ArrivalVehicle,
    BookingFloorAssignment,
    RawBookingAggregate,
    build_floor_target_breakdowns,
    build_arrival_vehicles,
    build_status_pallet_breakdowns,
    normalize_raw_sheet_booking,
    normalize_sequence_booking,
)
from Modules.Excel.MilkrunExcelImporter import ExcelImportCancelled, ExcelImportError


class ArrivalSequenceReaderTests(unittest.TestCase):
    @staticmethod
    def summary_values():
        return (
            (1, 2, 3, None, 10, 20, None, 4, 5, 6),
            (None,) * 10,
            (7, 8, 9, None, 30, 40, None, 11, 12, 13),
            (14, 15, 16, None, 100, 200, None, 17, 18, 19),
        )

    @staticmethod
    def detail_values():
        first = [None] * 22
        first[0] = "MBN123"
        first[14] = 99
        first[18] = "1F"
        first[21] = 99
        second = [None] * 22
        second[0] = "tbn00456"
        second[14] = 8
        return (tuple(first), tuple(second))

    @classmethod
    def create_workbook(
        cls,
        path: Path,
        *,
        include_sequence: bool = True,
        include_truck_floor: bool = True,
        include_milkrun_floor: bool = True,
    ) -> None:
        workbook = Workbook()
        sequence = workbook.active
        sequence.title = "입차순번" if include_sequence else "기타"
        if include_sequence:
            for row_offset, row in enumerate(cls.summary_values(), start=8):
                for column_offset, value in enumerate(row, start=37):
                    sequence.cell(row=row_offset, column=column_offset, value=value)
            for row_offset, row in enumerate(cls.detail_values(), start=18):
                for column_offset, value in enumerate(row, start=28):
                    sequence.cell(row=row_offset, column=column_offset, value=value)
        if include_truck_floor:
            truck = workbook.create_sheet("Raw_트럭")
            truck["B2"] = "1F"
            truck["C2"] = 456
        if include_milkrun_floor:
            milkrun = workbook.create_sheet("Raw_밀크런")
            milkrun["B2"] = "2F"
            milkrun["C2"] = 123
        workbook.save(path)
        workbook.close()

    def test_normalizes_sequence_reservation_prefixes(self) -> None:
        self.assertEqual(normalize_sequence_booking("MBN123"), ("M123", "milkrun"))
        self.assertEqual(normalize_sequence_booking(" mbn000123 "), ("M123", "milkrun"))
        self.assertEqual(normalize_sequence_booking("tbn00123"), ("T123", "truck"))
        self.assertEqual(normalize_sequence_booking("TBN0000456"), ("T456", "truck"))
        self.assertEqual(normalize_sequence_booking("M123"), ("", ""))
        self.assertEqual(normalize_raw_sheet_booking(123, prefix="M"), "M123")
        self.assertEqual(normalize_raw_sheet_booking("00123", prefix="T"), "T123")
        self.assertEqual(normalize_raw_sheet_booking("T123", prefix="T"), "T123")

    def test_reads_xlsm_memory_snapshot_with_loader_options_and_source_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            expected_modified = 1_700_000_000
            os.utime(path, (expected_modified, expected_modified))
            loader_calls: list[tuple[object, dict[str, object]]] = []
            loaded_snapshot_streams: list[object] = []

            def loader(snapshot_stream, **kwargs):
                self.assertTrue(snapshot_stream.readable())
                self.assertEqual(snapshot_stream.tell(), 0)
                loader_calls.append((snapshot_stream, kwargs))
                loaded_snapshot_streams.append(snapshot_stream)
                return load_workbook(snapshot_stream, **kwargs)

            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            )

            result = reader.read(path)

            self.assertEqual(result.summary.floor_targets[2], ("100", "200"))
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])
            self.assertEqual(result.entries[0].excel_row, 18)
            self.assertEqual(result.entries[0].unloading_floor, "1F")
            self.assertEqual(
                [(entry.booking_key, entry.floor) for entry in result.floor_assignments],
                [("T456", "1F"), ("M123", "2F")],
            )
            self.assertEqual(result.workbook, path.resolve())
            self.assertEqual(
                result.source_modified_at,
                datetime.fromtimestamp(path.stat().st_mtime),
            )
            self.assertEqual(len(loader_calls), 1)
            self.assertEqual(
                loader_calls[0][1],
                {
                    "read_only": True,
                    "data_only": True,
                    "keep_vba": False,
                    "keep_links": False,
                },
            )
            self.assertTrue(loaded_snapshot_streams[0].closed)

    def test_reads_saved_values_while_source_file_is_already_open(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)

            with path.open("rb") as existing_read_handle:
                result = ArrivalSequenceReader(sleep=lambda _delay: None).read(path)
                self.assertFalse(existing_read_handle.closed)

            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])

    def test_first_bad_zip_file_is_retried_and_then_succeeds(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            loader_calls = 0
            sleeps: list[float] = []
            logs: list[str] = []

            def loader(snapshot_path, **kwargs):
                nonlocal loader_calls
                loader_calls += 1
                if loader_calls == 1:
                    raise zipfile.BadZipFile("OneDrive is replacing the package")
                return load_workbook(snapshot_path, **kwargs)

            reader = ArrivalSequenceReader(
                log=logs.append,
                workbook_loader=loader,
                sleep=sleeps.append,
            )

            result = reader.read(path)

            self.assertEqual(loader_calls, 2)
            self.assertEqual(sleeps, [ArrivalSequenceReader._SNAPSHOT_RETRY_BASE_DELAY_SECONDS])
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])
            self.assertTrue(any("다시 읽습니다" in message for message in logs))

    def test_source_change_during_capture_is_retried(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            sleeps: list[float] = []
            reader = ArrivalSequenceReader(sleep=sleeps.append)
            original_signature = reader._source_signature
            signature_calls = 0

            def unstable_first_capture(stat_result):
                nonlocal signature_calls
                signature_calls += 1
                signature = original_signature(stat_result)
                if signature_calls == 4:
                    return (*signature[:-1], signature[-1] + 1)
                return signature

            with mock.patch.object(
                reader,
                "_source_signature",
                side_effect=unstable_first_capture,
            ):
                result = reader.read(path)

            self.assertEqual(signature_calls, 8)
            self.assertEqual(sleeps, [reader._SNAPSHOT_RETRY_BASE_DELAY_SECONDS])
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])

    def test_large_snapshot_rolls_from_memory_to_system_temp_storage(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            rolled: list[bool] = []

            def loader(snapshot_stream, **kwargs):
                rolled.append(bool(getattr(snapshot_stream, "_rolled", False)))
                return load_workbook(snapshot_stream, **kwargs)

            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            )
            reader._SNAPSHOT_MEMORY_LIMIT_BYTES = 1

            reader.read(path)

            self.assertEqual(rolled, [True])

    def test_retry_exhaustion_returns_sync_error(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            loader = mock.Mock(side_effect=zipfile.BadZipFile("still syncing"))
            sleeps: list[float] = []
            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=sleeps.append,
            )
            reader._SNAPSHOT_MAX_ATTEMPTS = 2

            with self.assertRaisesRegex(ArrivalSequenceError, "동기화된 Excel 저장본"):
                reader.read(path)

            self.assertEqual(loader.call_count, 2)
            self.assertEqual(sleeps, [reader._SNAPSHOT_RETRY_BASE_DELAY_SECONDS])

    def test_openpyxl_wrapped_xml_error_is_retried(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            loader_calls = 0

            def loader(snapshot_stream, **kwargs):
                nonlocal loader_calls
                loader_calls += 1
                if loader_calls == 1:
                    raise ValueError("Unable to read workbook: invalid XML")
                return load_workbook(snapshot_stream, **kwargs)

            result = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            ).read(path)

            self.assertEqual(loader_calls, 2)
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])

    def test_zip_compression_error_is_retried(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            loader_calls = 0

            def loader(snapshot_stream, **kwargs):
                nonlocal loader_calls
                loader_calls += 1
                if loader_calls == 1:
                    raise zlib.error("incomplete compressed stream")
                return load_workbook(snapshot_stream, **kwargs)

            result = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            ).read(path)

            self.assertEqual(loader_calls, 2)
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])

    def test_permanent_permission_error_is_not_retried(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            sleeps: list[float] = []
            reader = ArrivalSequenceReader(sleep=sleeps.append)

            with (
                mock.patch.object(
                    reader,
                    "_copy_stable_snapshot",
                    side_effect=PermissionError(13, "access denied"),
                ),
                self.assertRaisesRegex(ArrivalSequenceError, "파일 권한"),
            ):
                reader.read(path)

            self.assertEqual(sleeps, [])

    def test_missing_required_sheet_fails_without_retry(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path, include_milkrun_floor=False)
            loader_calls = 0
            sleeps: list[float] = []

            def loader(snapshot_path, **kwargs):
                nonlocal loader_calls
                loader_calls += 1
                return load_workbook(snapshot_path, **kwargs)

            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=sleeps.append,
            )

            with self.assertRaisesRegex(ArrivalSequenceError, "Raw_밀크런"):
                reader.read(path)

            self.assertEqual(loader_calls, 1)
            self.assertEqual(sleeps, [])

    def test_all_empty_summary_values_are_allowed(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            workbook = load_workbook(path)
            sequence = workbook["입차순번"]
            for row in range(8, 12):
                for column in range(37, 47):
                    sequence.cell(row=row, column=column).value = None
            workbook.save(path)
            workbook.close()

            result = ArrivalSequenceReader(sleep=lambda _delay: None).read(path)

            self.assertEqual(result.summary.departure[0], ("", "", ""))
            self.assertEqual([entry.booking_key for entry in result.entries], ["M123", "T456"])

    def test_xlsb_is_rejected_before_loading(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsb"
            path.write_bytes(b"not an OOXML package")
            loader = mock.Mock()
            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            )

            with self.assertRaisesRegex(ExcelImportError, "지원하지 않는 Excel 파일 형식"):
                reader.read(path)

            loader.assert_not_called()

    def test_cancelled_read_stops_before_snapshot_loading(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "SAN2_입고스케줄관리.xlsm"
            self.create_workbook(path)
            loader = mock.Mock()
            reader = ArrivalSequenceReader(
                workbook_loader=loader,
                sleep=lambda _delay: None,
            )

            with self.assertRaisesRegex(ExcelImportCancelled, "새로고침을 중지"):
                reader.read(path, cancel_requested=lambda: True)

            loader.assert_not_called()

    def test_current_raw_wins_and_previous_ap_aw_is_counted_once(self) -> None:
        snapshot = ArrivalSequenceSnapshot(
            workbook=Path("sample.xlsm"),
            sheet_name="입차순번",
            refreshed_at=__import__("datetime").datetime.now(),
            summary=ArrivalSummary((), (), ()),
            entries=(
                ArrivalSequenceEntry(18, "MBN123", "M123", "milkrun", "1F", 99, 99),
                ArrivalSequenceEntry(19, "tbn00456", "T456", "truck", "", 8, 8),
                ArrivalSequenceEntry(20, "MBN789", "M789", "milkrun", "2F", 7, 9),
            ),
        )
        raw = {
            "M123": RawBookingAggregate(
                "M123",
                ("거래처",),
                Decimal("3"),
                (("경량", Decimal("2")), ("고단", Decimal("1"))),
            )
        }

        vehicles = build_arrival_vehicles(snapshot, raw)

        self.assertEqual(vehicles[0].period, "금일")
        self.assertEqual(vehicles[0].pallet_count, Decimal("3"))
        self.assertEqual(vehicles[0].categories["고단"], Decimal("1"))
        self.assertEqual(vehicles[1].period, "전일")
        self.assertEqual(vehicles[1].pallet_count, Decimal("8"))
        self.assertEqual(vehicles[1].status, "외부대기")
        self.assertEqual(vehicles[2].pallet_count, Decimal("7"))
        self.assertIn("확인 필요", vehicles[2].note)

    def test_floor_target_breakdown_uses_excel_floor_and_app_raw_prefixes(self) -> None:
        snapshot = ArrivalSequenceSnapshot(
            workbook=Path("sample.xlsm"),
            sheet_name="입차순번",
            refreshed_at=__import__("datetime").datetime.now(),
            summary=ArrivalSummary((), (), ()),
            entries=(),
            floor_assignments=(
                BookingFloorAssignment("T123", "truck", "1F", "Raw_트럭", 2),
                BookingFloorAssignment("M456", "milkrun", "1F", "Raw_밀크런", 2),
                BookingFloorAssignment("T789", "truck", "2F", "Raw_트럭", 3),
            ),
        )
        raw = {
            "T123": RawBookingAggregate(
                "T123", ("트럭",), Decimal("2"), (("중량", Decimal("2")),)
            ),
            "M456": RawBookingAggregate(
                "M456", ("밀크런",), Decimal("3"), (("경량", Decimal("3")),)
            ),
            "T789": RawBookingAggregate(
                "T789", ("트럭2",), Decimal("4"), (("고단", Decimal("4")),)
            ),
        }

        first, second = build_floor_target_breakdowns(snapshot, raw)

        self.assertEqual((first.floor, first.truck_count, first.milkrun_count), ("1F", 1, 1))
        self.assertEqual(first.pallet_count, Decimal("5"))
        self.assertEqual(first.categories["경량"], Decimal("3"))
        self.assertEqual(first.categories["중량"], Decimal("2"))
        self.assertEqual((second.floor, second.truck_count), ("2F", 1))
        self.assertEqual(second.categories["고단"], Decimal("4"))
        self.assertEqual(first.unassigned_raw_bookings, ())

    def test_status_breakdowns_keep_vehicle_counts_separate_from_pallet_details(self) -> None:
        vehicles = (
            ArrivalVehicle(
                18,
                "M123",
                "milkrun",
                "거래처 A",
                "금일",
                "외부대기",
                "1F",
                Decimal("3"),
                (("경량", Decimal("2")), ("고단", Decimal("1"))),
            ),
            ArrivalVehicle(
                19,
                "T456",
                "truck",
                "거래처 B",
                "금일",
                "출차",
                "2F",
                Decimal("4"),
                (("중량", Decimal("4")),),
            ),
            ArrivalVehicle(
                20,
                "M789",
                "milkrun",
                "",
                "전일",
                "외부대기",
                "",
                Decimal("7"),
                (),
            ),
        )

        waiting = build_status_pallet_breakdowns(vehicles, status="외부대기")
        departure = build_status_pallet_breakdowns(vehicles, status="출차")

        self.assertEqual(waiting[0].label, "1F")
        self.assertEqual(waiting[0].pallet_count, Decimal("3"))
        self.assertEqual(waiting[0].categories["경량"], Decimal("2"))
        self.assertEqual(waiting[2].label, "전일자")
        self.assertEqual(waiting[2].pallet_count, Decimal("7"))
        self.assertEqual(departure[1].pallet_count, Decimal("4"))
        self.assertEqual(departure[1].categories["중량"], Decimal("4"))


if __name__ == "__main__":
    unittest.main()
