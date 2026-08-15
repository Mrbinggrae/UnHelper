from __future__ import annotations

import os
import re
import tempfile
import time
import zlib
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, LargeZipFile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO, Callable, Mapping

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from Modules.Excel.MilkrunExcelImporter import (
    ExcelImportCancelled,
    ExcelImportError,
    MilkrunExcelImporter,
)


LogCallback = Callable[[str], None]


class ArrivalSequenceError(ExcelImportError):
    """Raised when the linked arrival-sequence sheet cannot be read safely."""


class _WorkbookSnapshotChangedError(OSError):
    """Raised when OneDrive replaces or updates the source during capture."""


@dataclass(frozen=True)
class ArrivalSummary:
    departure: tuple[tuple[str, str, str], ...]
    outside_waiting: tuple[tuple[str, str, str], ...]
    floor_targets: tuple[tuple[str, str], ...]


@dataclass(frozen=True)
class ArrivalSequenceEntry:
    excel_row: int
    raw_reservation: str
    booking_key: str
    booking_type: str
    unloading_floor: str
    previous_ap_value: Any = None
    previous_aw_value: Any = None


@dataclass(frozen=True)
class BookingFloorAssignment:
    booking_key: str
    booking_type: str
    floor: str
    source_sheet: str
    source_row: int
    note: str = ""


@dataclass(frozen=True)
class ArrivalSequenceSnapshot:
    workbook: Path
    sheet_name: str
    refreshed_at: datetime
    summary: ArrivalSummary
    entries: tuple[ArrivalSequenceEntry, ...]
    floor_assignments: tuple[BookingFloorAssignment, ...] = ()
    source_modified_at: datetime | None = None


@dataclass(frozen=True)
class RawBookingAggregate:
    booking_key: str
    vendor_names: tuple[str, ...]
    pallet_count: Decimal
    category_pallets: tuple[tuple[str, Decimal], ...]
    missing_pallet_rows: int = 0

    @property
    def categories(self) -> dict[str, Decimal]:
        return dict(self.category_pallets)


@dataclass(frozen=True)
class ArrivalVehicle:
    excel_row: int
    booking_key: str
    booking_type: str
    vendor_name: str
    period: str
    status: str
    floor: str
    pallet_count: Decimal | None
    category_pallets: tuple[tuple[str, Decimal], ...]
    note: str = ""

    @property
    def categories(self) -> dict[str, Decimal]:
        return dict(self.category_pallets)


@dataclass(frozen=True)
class FloorTargetBreakdown:
    floor: str
    truck_count: int
    milkrun_count: int
    pallet_count: Decimal
    category_pallets: tuple[tuple[str, Decimal], ...]
    missing_pallet_rows: int = 0
    unmapped_bookings: tuple[str, ...] = ()
    unassigned_raw_bookings: tuple[str, ...] = ()

    @property
    def categories(self) -> dict[str, Decimal]:
        return dict(self.category_pallets)


@dataclass(frozen=True)
class ArrivalPalletBreakdown:
    label: str
    pallet_count: Decimal
    category_pallets: tuple[tuple[str, Decimal], ...]
    missing_pallet_vehicles: int = 0
    unmapped_bookings: tuple[str, ...] = ()
    notes: tuple[str, ...] = ()

    @property
    def categories(self) -> dict[str, Decimal]:
        return dict(self.category_pallets)


_MILKRUN_SEQUENCE_PATTERN = re.compile(r"^MBN\s*0*(\d+)\s*$", re.IGNORECASE)
_TRUCK_SEQUENCE_PATTERN = re.compile(r"^TBN00\s*0*(\d+)\s*$", re.IGNORECASE)
_FLOOR_PATTERN = re.compile(r"\b([12])\s*F\b", re.IGNORECASE)


def normalize_sequence_booking(value: Any) -> tuple[str, str]:
    """Map the sequence sheet's MBN/tbn00 values to RAW M/T keys."""

    candidate = re.sub(r"\s+", "", str(value or "")).strip()
    if not candidate:
        return "", ""
    match = _MILKRUN_SEQUENCE_PATTERN.fullmatch(candidate)
    if match:
        return f"M{match.group(1)}", "milkrun"
    match = _TRUCK_SEQUENCE_PATTERN.fullmatch(candidate)
    if match:
        return f"T{match.group(1)}", "truck"
    return "", ""


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    candidate = str(value).strip().replace(",", "")
    if not candidate:
        return None
    try:
        parsed = Decimal(candidate)
    except InvalidOperation:
        return None
    if not parsed.is_finite() or parsed < 0:
        return None
    return parsed


def normalize_raw_sheet_booking(value: Any, *, prefix: str) -> str:
    expected_prefix = str(prefix or "").strip().upper()
    if expected_prefix not in {"M", "T"} or value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        candidate = str(value)
    elif isinstance(value, float):
        candidate = str(int(value)) if value.is_integer() else str(value)
    else:
        candidate = str(value).strip().replace(",", "")
    candidate = re.sub(r"\s+", "", candidate).upper()
    if candidate.startswith(expected_prefix):
        candidate = candidate[1:]
    if candidate.endswith(".0") and candidate[:-2].isdigit():
        candidate = candidate[:-2]
    if not candidate.isdigit():
        return ""
    digits = candidate.lstrip("0") or "0"
    return f"{expected_prefix}{digits}"


def build_floor_target_breakdowns(
    snapshot: ArrivalSequenceSnapshot,
    raw_bookings: Mapping[str, RawBookingAggregate],
) -> tuple[FloorTargetBreakdown, ...]:
    """Aggregate current RAW vehicles by their B:C floor assignment."""

    categories = ("경량", "중량", "고단", "양곡", "?")
    states = {
        "1F": {
            "truck": 0,
            "milkrun": 0,
            "pallets": Decimal("0"),
            "categories": {category: Decimal("0") for category in categories},
            "missing": 0,
            "unmapped": [],
        },
        "2F": {
            "truck": 0,
            "milkrun": 0,
            "pallets": Decimal("0"),
            "categories": {category: Decimal("0") for category in categories},
            "missing": 0,
            "unmapped": [],
        },
    }
    seen: set[str] = set()
    assigned_raw_keys: set[str] = set()
    for assignment in snapshot.floor_assignments:
        if assignment.booking_key in seen or assignment.floor not in states:
            continue
        seen.add(assignment.booking_key)
        aggregate = raw_bookings.get(assignment.booking_key)
        state = states[assignment.floor]
        if aggregate is None:
            state["unmapped"].append(assignment.booking_key)
            continue
        assigned_raw_keys.add(assignment.booking_key)
        state[assignment.booking_type] += 1
        state["pallets"] += aggregate.pallet_count
        state["missing"] += aggregate.missing_pallet_rows
        for category, value in aggregate.category_pallets:
            if category in state["categories"]:
                state["categories"][category] += value

    unassigned_raw = tuple(sorted(set(raw_bookings) - assigned_raw_keys))
    return tuple(
        FloorTargetBreakdown(
            floor=floor,
            truck_count=int(state["truck"]),
            milkrun_count=int(state["milkrun"]),
            pallet_count=state["pallets"],
            category_pallets=tuple(
                (category, state["categories"][category])
                for category in categories
                if state["categories"][category] != 0
            ),
            missing_pallet_rows=int(state["missing"]),
            unmapped_bookings=tuple(state["unmapped"]),
            unassigned_raw_bookings=unassigned_raw,
        )
        for floor, state in states.items()
    )


def build_arrival_vehicles(
    snapshot: ArrivalSequenceSnapshot,
    raw_bookings: Mapping[str, RawBookingAggregate],
) -> tuple[ArrivalVehicle, ...]:
    """Combine live Excel status with current-day RAW pallet classifications."""

    vehicles: list[ArrivalVehicle] = []
    floor_assignments = {
        assignment.booking_key: assignment
        for assignment in snapshot.floor_assignments
    }
    for entry in snapshot.entries:
        floor_match = _FLOOR_PATTERN.search(entry.unloading_floor)
        assignment = floor_assignments.get(entry.booking_key)
        floor = (
            f"{floor_match.group(1)}F"
            if floor_match
            else (assignment.floor if assignment is not None else "")
        )
        status = "출차" if entry.unloading_floor.strip() else "외부대기"
        current = raw_bookings.get(entry.booking_key)
        if current is not None:
            notes = [assignment.note] if assignment is not None and assignment.note else []
            if current.missing_pallet_rows:
                notes.append(
                    f"팔렛트 수 확인 필요 SKU {current.missing_pallet_rows}행"
                )
            vehicles.append(
                ArrivalVehicle(
                    excel_row=entry.excel_row,
                    booking_key=entry.booking_key,
                    booking_type=entry.booking_type,
                    vendor_name=" · ".join(current.vendor_names),
                    period="금일",
                    status=status,
                    floor=floor,
                    pallet_count=current.pallet_count,
                    category_pallets=current.category_pallets,
                    note=" · ".join(notes),
                )
            )
            continue

        ap_value = decimal_or_none(entry.previous_ap_value)
        aw_value = decimal_or_none(entry.previous_aw_value)
        previous_value = ap_value if ap_value is not None else aw_value
        notes = [assignment.note] if assignment is not None and assignment.note else []
        if ap_value is not None and aw_value is not None and ap_value != aw_value:
            notes.append(f"전일 팔렛트 값 확인 필요 (AP {ap_value} / AW {aw_value})")
        elif previous_value is None:
            notes.append("전일 팔렛트 수 미입력")
        vehicles.append(
            ArrivalVehicle(
                excel_row=entry.excel_row,
                booking_key=entry.booking_key,
                booking_type=entry.booking_type,
                vendor_name="",
                period="전일",
                status=status,
                floor=floor,
                pallet_count=previous_value,
                category_pallets=(),
                note=" · ".join(notes),
            )
        )
    return tuple(vehicles)


def build_status_pallet_breakdowns(
    vehicles: tuple[ArrivalVehicle, ...],
    *,
    status: str,
) -> tuple[ArrivalPalletBreakdown, ...]:
    """Summarize pallets below the Excel-provided vehicle-count cards."""

    categories = ("경량", "중량", "고단", "양곡", "?")
    states = {
        label: {
            "pallets": Decimal("0"),
            "categories": {category: Decimal("0") for category in categories},
            "missing": 0,
            "unmapped": [],
            "notes": [],
        }
        for label in ("1F", "2F", "전일자")
    }
    for vehicle in vehicles:
        if vehicle.status != status:
            continue
        label = "전일자" if vehicle.period == "전일" else vehicle.floor
        if label not in states:
            states["1F"]["unmapped"].append(vehicle.booking_key)
            continue
        state = states[label]
        if vehicle.pallet_count is None:
            state["missing"] += 1
        else:
            state["pallets"] += vehicle.pallet_count
        for category, value in vehicle.category_pallets:
            if category in state["categories"]:
                state["categories"][category] += value
        if vehicle.note:
            state["notes"].append(f"{vehicle.booking_key}: {vehicle.note}")

    return tuple(
        ArrivalPalletBreakdown(
            label=label,
            pallet_count=state["pallets"],
            category_pallets=tuple(
                (category, state["categories"][category])
                for category in categories
                if state["categories"][category] != 0
            ),
            missing_pallet_vehicles=int(state["missing"]),
            unmapped_bookings=tuple(state["unmapped"]),
            notes=tuple(state["notes"]),
        )
        for label, state in states.items()
    )


class ArrivalSequenceReader(MilkrunExcelImporter):
    """Read the synchronized workbook snapshot without starting Excel."""

    TARGET_SHEET = "입차순번"
    TARGET_EXTENSIONS = frozenset({".xlsx", ".xlsm"})
    FIRST_DETAIL_ROW = 18
    MAX_DETAIL_ROW = 5000
    MAX_RAW_ROW = 10000
    _SEQUENCE_FIRST_ROW = 8
    _SEQUENCE_FIRST_COLUMN = 28  # AB
    _SEQUENCE_LAST_COLUMN = 49  # AW
    _SUMMARY_FIRST_COLUMN_OFFSET = 9  # AK - AB
    _SUMMARY_COLUMN_COUNT = 10  # AK:AT
    _SNAPSHOT_MAX_ATTEMPTS = 4
    _SNAPSHOT_COPY_CHUNK_BYTES = 2 * 1024 * 1024
    _SNAPSHOT_MEMORY_LIMIT_BYTES = 32 * 1024 * 1024
    _SNAPSHOT_RETRY_BASE_DELAY_SECONDS = 0.25

    def __init__(
        self,
        log: LogCallback | None = None,
        *,
        workbook_loader: Callable[..., Any] | None = None,
        sleep: Callable[[float], None] | None = None,
    ) -> None:
        super().__init__(log=log, reject_open_target=False)
        self._workbook_loader = workbook_loader or load_workbook
        self._sleep = sleep or time.sleep

    def read(
        self,
        workbook_path: str | Path,
        *,
        cancel_requested: Callable[[], bool] | None = None,
    ) -> ArrivalSequenceSnapshot:
        target_path = self.validate_target_path(workbook_path)
        last_error: Exception | None = None
        for attempt in range(1, self._SNAPSHOT_MAX_ATTEMPTS + 1):
            self._raise_if_cancelled(cancel_requested)
            try:
                snapshot = self._read_snapshot_once(
                    target_path,
                    cancel_requested=cancel_requested,
                )
            except ExcelImportCancelled:
                raise
            except ArrivalSequenceError:
                raise
            except (InvalidFileException, LargeZipFile) as exc:
                raise ArrivalSequenceError(
                    "입차순번 읽기는 암호화되지 않은 .xlsx 또는 .xlsm 저장본만 "
                    f"지원합니다.\n{exc}"
                ) from exc
            except Exception as exc:
                if not self._is_retryable_snapshot_error(exc):
                    if isinstance(exc, OSError):
                        raise ArrivalSequenceError(
                            "연결된 Excel 저장본에 접근하지 못했습니다. "
                            "파일 권한과 디스크 여유 공간을 확인해 주세요.\n"
                            f"{exc}"
                        ) from exc
                    raise
                last_error = exc
                if attempt >= self._SNAPSHOT_MAX_ATTEMPTS:
                    break
                self.log(
                    "Excel 또는 OneDrive가 저장본을 갱신 중입니다. "
                    f"잠시 후 다시 읽습니다 ({attempt}/{self._SNAPSHOT_MAX_ATTEMPTS})."
                )
                self._wait_for_retry(attempt, cancel_requested)
                continue

            self.log(
                "Excel을 실행하지 않고 마지막 저장·동기화 값에서 "
                f"차량 {len(snapshot.entries)}건을 읽었습니다."
            )
            return snapshot

        raise ArrivalSequenceError(
            "동기화된 Excel 저장본의 '입차순번' 값을 읽지 못했습니다. "
            "웹 Excel의 저장 완료와 OneDrive 동기화 상태를 확인한 뒤 다시 시도해 주세요.\n"
            f"{last_error}"
        ) from last_error

    def _read_snapshot_once(
        self,
        target_path: Path,
        *,
        cancel_requested: Callable[[], bool] | None,
    ) -> ArrivalSequenceSnapshot:
        with tempfile.SpooledTemporaryFile(
            max_size=self._SNAPSHOT_MEMORY_LIMIT_BYTES,
            mode="w+b",
        ) as snapshot_stream:
            source_stat = self._copy_stable_snapshot(
                target_path,
                snapshot_stream,
                cancel_requested=cancel_requested,
            )
            snapshot_stream.seek(0)
            workbook = self._workbook_loader(
                snapshot_stream,
                read_only=True,
                data_only=True,
                keep_vba=False,
                keep_links=False,
            )
            try:
                sequence_sheet = self._required_sheet(workbook, self.TARGET_SHEET)
                sequence_values = self._read_rows(
                    sequence_sheet,
                    min_row=self._SEQUENCE_FIRST_ROW,
                    max_row=self.MAX_DETAIL_ROW,
                    min_column=self._SEQUENCE_FIRST_COLUMN,
                    max_column=self._SEQUENCE_LAST_COLUMN,
                    cancel_requested=cancel_requested,
                )
                summary_start = self._SUMMARY_FIRST_COLUMN_OFFSET
                summary_end = summary_start + self._SUMMARY_COLUMN_COUNT
                summary_values = tuple(
                    tuple(row[summary_start:summary_end])
                    for row in sequence_values[:4]
                )
                detail_offset = self.FIRST_DETAIL_ROW - self._SEQUENCE_FIRST_ROW
                detail_values = sequence_values[detail_offset:]

                floor_values: dict[str, tuple[tuple[Any, ...], ...]] = {}
                for sheet_name, _booking_type, _prefix in self._floor_sheet_specs():
                    floor_sheet = self._required_sheet(workbook, sheet_name)
                    floor_values[sheet_name] = self._read_rows(
                        floor_sheet,
                        min_row=2,
                        max_row=self.MAX_RAW_ROW,
                        min_column=2,
                        max_column=3,
                        cancel_requested=cancel_requested,
                    )
            finally:
                workbook.close()

        return ArrivalSequenceSnapshot(
            workbook=target_path,
            sheet_name=self.TARGET_SHEET,
            refreshed_at=datetime.now(),
            summary=self._parse_summary(summary_values),
            entries=self._parse_entries(detail_values),
            floor_assignments=self._parse_floor_assignments(floor_values),
            source_modified_at=datetime.fromtimestamp(source_stat.st_mtime),
        )

    def _copy_stable_snapshot(
        self,
        source_path: Path,
        snapshot_stream: BinaryIO,
        *,
        cancel_requested: Callable[[], bool] | None,
    ) -> os.stat_result:
        before = source_path.stat()
        with source_path.open("rb") as source:
            opened_before = os.fstat(source.fileno())
            while True:
                self._raise_if_cancelled(cancel_requested)
                chunk = source.read(self._SNAPSHOT_COPY_CHUNK_BYTES)
                if not chunk:
                    break
                snapshot_stream.write(chunk)
            opened_after = os.fstat(source.fileno())
        snapshot_stream.flush()
        after = source_path.stat()
        signatures = {
            self._source_signature(stat_result)
            for stat_result in (before, opened_before, opened_after, after)
        }
        if len(signatures) != 1:
            raise _WorkbookSnapshotChangedError(
                "OneDrive 동기화 중 원본 파일이 변경되었습니다."
            )
        return after

    @staticmethod
    def _source_signature(stat_result: os.stat_result) -> tuple[int, int, int, int]:
        return (
            int(stat_result.st_dev),
            int(stat_result.st_ino),
            int(stat_result.st_size),
            int(stat_result.st_mtime_ns),
        )

    @staticmethod
    def _required_sheet(workbook: Any, sheet_name: str) -> Any:
        try:
            return workbook[sheet_name]
        except KeyError as exc:
            raise ArrivalSequenceError(
                f"연결된 Excel 파일에 '{sheet_name}' 시트가 없습니다."
            ) from exc

    @staticmethod
    def _is_retryable_snapshot_error(exc: Exception) -> bool:
        if isinstance(
            exc,
            (
                _WorkbookSnapshotChangedError,
                BadZipFile,
                EOFError,
                ParseError,
                zlib.error,
                KeyError,
                FileNotFoundError,
                TimeoutError,
            ),
        ):
            return True
        if isinstance(exc, ValueError):
            return "Unable to read workbook" in str(exc)
        if isinstance(exc, OSError):
            return getattr(exc, "winerror", None) in {32, 33, 995}
        return False

    def _read_rows(
        self,
        sheet: Any,
        *,
        min_row: int,
        max_row: int,
        min_column: int,
        max_column: int,
        cancel_requested: Callable[[], bool] | None,
    ) -> tuple[tuple[Any, ...], ...]:
        sheet_max_row = int(getattr(sheet, "max_row", max_row) or max_row)
        bounded_max_row = min(max_row, sheet_max_row)
        if bounded_max_row < min_row:
            return ()

        rows: list[tuple[Any, ...]] = []
        for offset, row in enumerate(
            sheet.iter_rows(
                min_row=min_row,
                max_row=bounded_max_row,
                min_col=min_column,
                max_col=max_column,
                values_only=True,
            )
        ):
            if offset % 128 == 0:
                self._raise_if_cancelled(cancel_requested)
            rows.append(tuple(row))
        while rows and all(value in (None, "") for value in rows[-1]):
            rows.pop()
        return tuple(rows)

    @staticmethod
    def _raise_if_cancelled(
        cancel_requested: Callable[[], bool] | None,
    ) -> None:
        if cancel_requested is not None and cancel_requested():
            raise ExcelImportCancelled("사용자가 입차순번 새로고침을 중지했습니다.")

    def _wait_for_retry(
        self,
        attempt: int,
        cancel_requested: Callable[[], bool] | None,
    ) -> None:
        self._raise_if_cancelled(cancel_requested)
        self._sleep(self._SNAPSHOT_RETRY_BASE_DELAY_SECONDS * attempt)
        self._raise_if_cancelled(cancel_requested)

    @staticmethod
    def _floor_sheet_specs() -> tuple[tuple[str, str, str], ...]:
        return (
            ("Raw_트럭", "truck", "T"),
            ("Raw_밀크런", "milkrun", "M"),
        )

    def _parse_floor_assignments(
        self,
        values_by_sheet: Mapping[str, Any],
    ) -> tuple[BookingFloorAssignment, ...]:
        assignments: list[BookingFloorAssignment] = []
        for sheet_name, booking_type, prefix in self._floor_sheet_specs():
            for offset, row in enumerate(self._matrix(values_by_sheet.get(sheet_name))):
                floor_text = str(row[0] or "").strip() if len(row) > 0 else ""
                floor_match = _FLOOR_PATTERN.search(floor_text)
                booking_key = normalize_raw_sheet_booking(
                    row[1] if len(row) > 1 else None,
                    prefix=prefix,
                )
                if not booking_key:
                    continue
                assignments.append(
                    BookingFloorAssignment(
                        booking_key=booking_key,
                        booking_type=booking_type,
                        floor=f"{floor_match.group(1)}F" if floor_match else "",
                        source_sheet=sheet_name,
                        source_row=2 + offset,
                        note="" if floor_match else f"층 정보 확인 필요: {floor_text or '(빈값)'}",
                    )
                )
        return self._deduplicate_floor_assignments(assignments)

    @staticmethod
    def _deduplicate_floor_assignments(
        assignments: list[BookingFloorAssignment],
    ) -> tuple[BookingFloorAssignment, ...]:
        result: list[BookingFloorAssignment] = []
        positions: dict[str, int] = {}
        for assignment in assignments:
            position = positions.get(assignment.booking_key)
            if position is None:
                positions[assignment.booking_key] = len(result)
                result.append(assignment)
                continue
            previous = result[position]
            if not previous.floor and assignment.floor:
                result[position] = assignment
                continue
            if previous.floor and not assignment.floor:
                continue
            if previous.floor == assignment.floor:
                continue
            floors = "/".join(
                value for value in dict.fromkeys((previous.floor, assignment.floor)) if value
            ) or "미입력"
            result[position] = BookingFloorAssignment(
                booking_key=previous.booking_key,
                booking_type=previous.booking_type,
                floor="",
                source_sheet=previous.source_sheet,
                source_row=previous.source_row,
                note=f"층 정보 충돌: {floors}",
            )
        return tuple(result)

    @staticmethod
    def _matrix(values: Any) -> tuple[tuple[Any, ...], ...]:
        return MilkrunExcelImporter._to_matrix(values) if values not in (None, ()) else ()

    @classmethod
    def _parse_summary(cls, values: Any) -> ArrivalSummary:
        matrix = cls._matrix(values)

        def at(row: int, column: int) -> str:
            try:
                value = matrix[row][column]
            except IndexError:
                return ""
            if value is None:
                return ""
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        return ArrivalSummary(
            departure=(
                (at(0, 0), at(0, 1), at(0, 2)),
                (at(2, 0), at(2, 1), at(2, 2)),
                (at(3, 0), at(3, 1), at(3, 2)),
            ),
            outside_waiting=(
                (at(0, 7), at(0, 8), at(0, 9)),
                (at(2, 7), at(2, 8), at(2, 9)),
                (at(3, 7), at(3, 8), at(3, 9)),
            ),
            floor_targets=(
                (at(0, 4), at(0, 5)),
                (at(2, 4), at(2, 5)),
                (at(3, 4), at(3, 5)),
            ),
        )

    @classmethod
    def _parse_entries(cls, values: Any) -> tuple[ArrivalSequenceEntry, ...]:
        entries: list[ArrivalSequenceEntry] = []
        for offset, row in enumerate(cls._matrix(values)):
            raw_reservation = row[0] if len(row) > 0 else None
            booking_key, booking_type = normalize_sequence_booking(raw_reservation)
            if not booking_key:
                continue
            entries.append(
                ArrivalSequenceEntry(
                    excel_row=cls.FIRST_DETAIL_ROW + offset,
                    raw_reservation=str(raw_reservation or "").strip(),
                    booking_key=booking_key,
                    booking_type=booking_type,
                    previous_ap_value=row[14] if len(row) > 14 else None,
                    unloading_floor=str(row[18] or "").strip() if len(row) > 18 else "",
                    previous_aw_value=row[21] if len(row) > 21 else None,
                )
            )
        return tuple(entries)
